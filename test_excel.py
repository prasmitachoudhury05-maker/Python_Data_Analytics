import sys, traceback
from openpyxl import load_workbook

path = r"Advanced\quarter_1.xlsx"
try:
    wb = load_workbook(filename=path, read_only=True)
    print("Workbook loaded successfully. Sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(row)
        if i >= 5:
            break
except Exception as e:
    print("Error loading workbook:", e)
    traceback.print_exc()